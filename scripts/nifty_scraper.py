"""
NiftyIndices Benchmark Index Scraper
Source  : niftyindices.com  (internal POST endpoint used by the website)
Output  : benchmark_nifty.xlsx

How it works
------------
niftyindices.com has no public API, but their Historical Data page POSTs to:
  https://www.niftyindices.com/Backpage.aspx/getHistoricaldatatabletoString

The request requires:
  1. A valid session cookie obtained by a GET to the homepage first.
  2. A JSON body: {"cinfo": '{"name":"<INDEX>","startDate":"<DD-Mon-YYYY>","endDate":"<DD-Mon-YYYY>"}'}
  3. Specific headers (Content-Type, Referer, X-Requested-With).

The response is JSON whose "d" key contains an HTML table string.
We parse that with BeautifulSoup to extract the OHLC data.

Install
-------
    pip install requests beautifulsoup4 lxml openpyxl pandas

Usage
-----
    python scraper_nifty.py
"""

import json
import time
from io import StringIO

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────────────────────
FROM_DATE   = "31-Dec-2020"   # DD-Mon-YYYY  (format required by the endpoint)
TO_DATE     = "31-Jan-2026"
OUTPUT_FILE = "benchmark_nifty.xlsx"

# ── Index list ─────────────────────────────────────────────────────────────────
# Keys   : column headers in the output Excel
# Values : exact index name strings as used by niftyindices.com
#
# To find the correct string for any index, go to:
#   https://www.niftyindices.com/reports/historical-data
# open DevTools → Network, select an index, and inspect the POST payload.
#
BENCHMARKS = {
    "Nifty 50":            "NIFTY 50",
    "Nifty IT":            "NIFTY IT",
    "Nifty Healthcare":    "NIFTY HEALTHCARE INDEX",
    "Nifty Pharma":        "NIFTY PHARMA",
    "Nifty Bank":          "NIFTY BANK",
    "Nifty FMCG":          "NIFTY FMCG",
    "Nifty Midcap 150":    "NIFTY MIDCAP 150",
    "Nifty Smallcap 250":  "NIFTY SMALLCAP 250",
    "Nifty Financial Svcs":"NIFTY FINANCIAL SERVICES",
}

# ── API constants ──────────────────────────────────────────────────────────────
BASE_URL     = "https://www.niftyindices.com"
ENDPOINT     = f"{BASE_URL}/Backpage.aspx/getHistoricaldatatabletoString"
USER_AGENT   = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

# ── Excel styling ──────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", start_color="155724")   # dark green for Nifty
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="D4EDDA")
STD_FONT = Font(name="Arial", size=10)
CENTER   = Alignment(horizontal="center")
RIGHT    = Alignment(horizontal="right")


# ── Session setup ──────────────────────────────────────────────────────────────

def make_session() -> requests.Session:
    """
    Create a requests Session, warm it up by GETting the homepage
    (this sets the required cookies), and return it ready to POST.
    """
    session = requests.Session()
    session.headers.update({
        "User-Agent":      USER_AGENT,
        "Accept-Language": "en-US,en;q=0.9",
    })
    print("  Warming up session (fetching cookies)…")
    resp = session.get(BASE_URL, timeout=30)
    resp.raise_for_status()
    return session


# ── Fetch helpers ──────────────────────────────────────────────────────────────

def post_index(session: requests.Session, index_name: str) -> list[dict]:
    """
    POST to the NiftyIndices endpoint and return a list of row dicts.
    Returns [] on failure.
    """
    headers = {
        "Content-Type":     "application/json; charset=UTF-8",
        "Accept":           "application/json, text/javascript, */*; q=0.01",
        "Referer":          f"{BASE_URL}/reports/historical-data",
        "X-Requested-With": "XMLHttpRequest",
        "Origin":           BASE_URL,
    }
    # The cinfo value must be a JSON-encoded string (double-serialised)
    cinfo = json.dumps({
        "name":      index_name,
        "startDate": FROM_DATE,
        "endDate":   TO_DATE,
    })
    payload = {"cinfo": cinfo}

    resp = session.post(ENDPOINT, json=payload, headers=headers, timeout=60)
    resp.raise_for_status()

    outer = resp.json()
    if "d" not in outer:
        print(f"     ⚠  Unexpected response: {outer}")
        return []

    # "d" is an HTML table string; parse it
    html = outer["d"]
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table")
    if not table:
        return []

    rows = []
    headers_row = [th.get_text(strip=True) for th in table.find_all("th")]
    for tr in table.find_all("tr")[1:]:  # skip header row
        cells = [td.get_text(strip=True) for td in tr.find_all("td")]
        if cells:
            rows.append(dict(zip(headers_row, cells)))
    return rows


def fetch_index_eom(session: requests.Session, index_name: str, label: str) -> pd.Series:
    """
    Fetch daily data for one index and return an end-of-month pd.Series.
    """
    print(f"  → {label}  ({index_name})")
    try:
        rows = post_index(session, index_name)
        if not rows:
            print(f"     ⚠  No rows returned")
            return pd.Series(dtype=float, name=label)

        df = pd.DataFrame(rows)
        df.columns = [c.strip().lower() for c in df.columns]

        # Identify date and close columns flexibly
        date_col  = next(c for c in df.columns if "date" in c)
        close_col = next(c for c in df.columns if "close" in c)

        df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors="coerce")
        close = (
            df.dropna(subset=[date_col])
              .set_index(date_col)
              .sort_index()[close_col]
        )
        # NiftyIndices returns numbers with commas; strip them
        close = pd.to_numeric(
            close.astype(str).str.replace(",", "", regex=False),
            errors="coerce",
        )

        eom = close.resample("ME").last()
        eom.name = label
        return eom

    except StopIteration:
        print(f"     ✗  Cannot identify date/close column")
        return pd.Series(dtype=float, name=label)
    except Exception as exc:
        print(f"     ✗  {exc}")
        return pd.Series(dtype=float, name=label)


def build_df(session: requests.Session) -> pd.DataFrame:
    series = []
    for label, index_name in BENCHMARKS.items():
        series.append(fetch_index_eom(session, index_name, label))
        time.sleep(1.0)   # be polite; NiftyIndices rate-limits aggressively

    df = pd.concat(series, axis=1)
    df.index = pd.to_datetime(df.index, errors="coerce")
    df = df[~df.index.isna()].sort_index()
    df.index = df.index.strftime("%d-%b-%Y")
    df.index.name = "Month End Date"
    return df


# ── Excel writer ───────────────────────────────────────────────────────────────

def style_ws(ws, df: pd.DataFrame):
    headers = ["Month End Date"] + list(df.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER

    for ri, (date_str, row) in enumerate(df.iterrows(), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        dc = ws.cell(row=ri, column=1, value=date_str)
        dc.font = STD_FONT; dc.alignment = CENTER
        if fill: dc.fill = fill

        for ci, val in enumerate(row, 2):
            cell = ws.cell(
                row=ri, column=ci,
                value=round(float(val), 4) if pd.notna(val) else "N/A",
            )
            cell.font = STD_FONT
            cell.alignment = RIGHT
            cell.number_format = "#,##0.0000"
            if fill: cell.fill = fill

    ws.column_dimensions["A"].width = 16
    for ci in range(2, len(df.columns) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 26
    ws.freeze_panes = "A2"


def save_excel(df: pd.DataFrame, path: str):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Nifty Benchmarks")
    style_ws(ws, df)
    wb.save(path)
    print(f"\n✅  Saved → {path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  NiftyIndices Scraper  |  POST endpoint")
    print(f"  Period : {FROM_DATE}  →  {TO_DATE}")
    print("=" * 60)

    session = make_session()
    df = build_df(session)
    session.close()

    if df.empty:
        print("\n❌  No data fetched. Check index names or connectivity.")
        return

    empty = [c for c in df.columns if df[c].isna().all()]
    if empty:
        print(f"\n⚠  Entirely empty columns: {', '.join(empty)}")
        print("   Verify the index name string matches exactly what the site uses.")
        print("   Tip: Open DevTools on niftyindices.com/reports/historical-data,")
        print("        select an index, and inspect the POST payload's 'cinfo.name'.")

    print(f"\n  Rows : {len(df)}   Columns : {len(df.columns)}")
    print(df.tail(3))
    save_excel(df, OUTPUT_FILE)


if __name__ == "__main__":
    main()