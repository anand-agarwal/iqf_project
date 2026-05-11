"""
Fetch end-of-month NAV/Index data from Yahoo Finance for Indian Mutual Funds
and BSE benchmark indices.

Date range: 01 Jan 2016 – 31 Jan 2020
Output: nav_data.xlsx  (one sheet per fund category + one "Benchmarks" sheet
                         + a combined "All Funds" sheet)

Usage:
    pip install yfinance pandas openpyxl
    python data_collection.py
"""

import time
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Date window ────────────────────────────────────────────────────────────────
# yfinance 'start' is inclusive, 'end' is exclusive.
# We fetch a slightly wider window and then trim to the exact EOM range.
START      = "2020-12-01"   # fetch from here …
END        = "2026-02-01"   # … up to (but not including) this date
TRIM_START = "2020-12-31"   # first EOM we want in the output
TRIM_END   = "2026-01-31"   # last  EOM we want in the output

# ── Benchmark universe ─────────────────────────────────────────────────────────
# These are the S&P BSE sector indices available on Yahoo Finance.
# Add / remove as needed.
BENCHMARKS = [
    ("BSE Teck Index",        "BSE-TECK.BO"),   # Technology / Media / Telecom
    ("BSE Healthcare Index",  "BSE-HC.BO"),      # Health Care
    ("BSE IT Index",          "BSE-IT.BO"),      # Pure Information Technology
    ("BSE FMCG Index",        "BSE-FMCG.BO"),   # Fast-Moving Consumer Goods
    ("BSE Bankex",            "BSE-BANK.BO"),    # Banking
    ("BSE Capital Goods",     "BSE-CG.BO"),      # Capital Goods
    ("BSE Smallcap",          "BSE-SMLCAP.BO"), # Small Cap broad index
    ("BSE Midcap",            "BSE-MIDCAP.BO"), # Mid Cap broad index
    ("BSE Sensex",            "^BSESN"),          # Nifty equivalent (broad market)
]

# ── Fund universe ──────────────────────────────────────────────────────────────
FUNDS = {
    # Uncomment the categories you need:

    # "Small Cap": [
    #     ("DSP Small Cap Dir Gr",              "0P0000XW24.BO"),
    #     ("Quant Small Cap Dir Gr",            "0P0000XW4J.BO"),
    #     ("Axis Small Cap Fund Dir Gr",        "0P00011MAX.BO"),
    #     ("Kotak Small Cap Dir Gr",            "0P0000XV6I.BO"),
    #     ("Nippon India Small Cap Dir Gr",     "0P0000XVFY.BO"),
    # ],
    # "Mid Cap": [
    #     ("HDFC Mid Cap Dir Gr",               "0P0000XW8F.BO"),
    #     ("Motilal Oswal Midcap Dir Gr",       "0P00012ALS.BO"),
    #     ("SBI Midcap Dir Gr",                 "0P0000XVKO.BO"),
    #     ("Axis Midcap Dir Gr",                "0P0000XVUH.BO"),
    #     ("Kotak Midcap Dir Gr",               "0P0000XV5R.BO"),
    # ],
    # "Large Cap": [
    #     ("ICICI Prudential Large Cap Dir Gr", "0P0000XWAT.BO"),
    #     ("Mirae Asset Large Cap Dir Gr",      "0P0000XVA0.BO"),
    #     ("HDFC Large Cap Dir Gr",             "0P0000XW91.BO"),
    #     ("Kotak Large Cap Dir Gr",            "0P0000XV5I.BO"),
    # ],
    # "Contra": [
    #     ("SBI Contra Dir Gr",                 "0P0000XVJR.BO"),
    #     ("Invesco India Contra Dir Gr",       "0P0000XVGR.BO"),
    #     ("Kotak Contra Dir Gr",               "0P0000XV5Q.BO"),
    # ],
    # "IT Funds": [
    #     ("Franklin India Technology Dir Gr",  "0P0000XW5R.BO"),
    #     ("SBI Technology Opportunities Dir Gr","0P0000XVKP.BO"),
    #     ("Tata Digital India Dir Gr",         "0P0001784G.BO"),
    #     ("HDFC Technology Dir Gr",            "0P0001RINJ.BO"),
    #     ("Kotak Technology Fund Dir Gr",      "0P0001SDWZ.BO"),
    # ],
    # "Pharma Funds": [
    #     ("SBI Healthcare Opportunities Dir Gr","0P0000XVL7.BO"),
    #     ("UTI Healthcare Dir Gr",             "0P0000XVU8.BO"),
    #     ("Mirae Asset Healthcare Dir Gr",     "0P0001DHEX.BO"),
    #     ("Nippon India Pharma Dir Gr",        "0P0000XVFK.BO"),
    #     ("Tata India Pharma & Healthcare Dir Gr","0P0001784W.BO"),
    # ],
    # "Banking Funds": [
    #     ("Invesco India Financial Serv Dir Gr","0P0000XVGN.BO"),
    #     ("SBI Banking & Financial Svcs Dir Gr","0P00015HLN.BO"),
    #     ("Nippon India Banking & Fin Services","0P0001BB28.BO"),
    #     ("Sundaram Fin Services Opp Dir Gr",  "0P0000XVM5.BO"),
    #     ("Mirae Asset Banking & Fin Svcs Dir Gr","0P0001L5SF.BO"),
    # ],
    # "Liquid Funds": [
    #     ("Bank of India Liquid Dir Gr",       "0P0000XVZO.BO"),
    #     ("Axis Liquid Dir Gr",                "0P0000XVUB.BO"),
    #     ("Canara Robeco Liquid Dir Gr",       "0P0000XW0R.BO"),
    #     ("Franklin India Liquid Fund",        "0P0001BA6D.BO"),
    #     ("DSP Liquidity Fund",                "0P00005V9P.BO"),
    # ],
}

# ── Data fetching ──────────────────────────────────────────────────────────────

def fetch_eom_series(ticker: str, name: str) -> pd.Series:
    """
    Download daily Close prices and return end-of-month values
    trimmed to [TRIM_START, TRIM_END].
    Works for both mutual fund NAV tickers (*.BO) and index tickers (^BSESN).
    """
    try:
        df = yf.download(ticker, start=START, end=END,
                         auto_adjust=True, progress=False)
        if df.empty:
            print(f"  ⚠  No data for {ticker}")
            return pd.Series(dtype=float, name=name)

        # Flatten MultiIndex columns if yfinance returns them
        close = df["Close"]
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]

        close.index = pd.to_datetime(close.index)

        # Last trading day of each calendar month
        eom = close.resample("ME").last()

        # Trim to the requested window
        eom = eom.loc[TRIM_START:TRIM_END]
        eom.name = name
        return eom

    except Exception as e:
        print(f"  ✗  Error fetching {ticker}: {e}")
        return pd.Series(dtype=float, name=name)


def build_df_from_list(fund_list: list) -> pd.DataFrame:
    """Fetch every (name, ticker) pair and combine into one DataFrame."""
    series_list = []
    for name, ticker in fund_list:
        print(f"  Fetching: {name} ({ticker})")
        s = fetch_eom_series(ticker, name)
        series_list.append(s)
        time.sleep(0.4)   # gentle rate-limiting

    if not series_list:
        return pd.DataFrame()

    df = pd.concat(series_list, axis=1)

    if df.empty:
        df.index = pd.Index([], name="Month End Date")
        return df

    df.index = pd.to_datetime(df.index, errors="coerce")
    df = df[~df.index.isna()].sort_index()
    df.index = df.index.strftime("%d-%b-%Y")
    df.index.name = "Month End Date"
    return df


# ── Excel styling ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="EBF3FB")
NORMAL_FONT = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center")
RIGHT       = Alignment(horizontal="right")


def style_sheet(ws, df: pd.DataFrame):
    """Apply consistent formatting to a worksheet."""
    # Header row
    for col_idx, col_name in enumerate(["Month End Date"] + list(df.columns), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    # Data rows
    for row_idx, (date_str, row) in enumerate(df.iterrows(), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None

        date_cell = ws.cell(row=row_idx, column=1, value=date_str)
        date_cell.font      = NORMAL_FONT
        date_cell.alignment = CENTER
        if fill:
            date_cell.fill = fill

        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(
                row=row_idx, column=col_idx,
                value=round(float(val), 4) if pd.notna(val) else "N/A"
            )
            cell.font          = NORMAL_FONT
            cell.alignment     = RIGHT
            cell.number_format = '#,##0.0000'
            if fill:
                cell.fill = fill

    # Column widths & freeze
    ws.column_dimensions["A"].width = 16
    for col_idx in range(2, len(df.columns) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 30
    ws.freeze_panes = "A2"


def write_excel(category_dfs: dict, benchmark_df: pd.DataFrame,
                all_df: pd.DataFrame, path: str):
    wb = Workbook()
    wb.remove(wb.active)   # drop default sheet

    # "All Funds" combined sheet (first tab)
    ws_all = wb.create_sheet(title="All Funds", index=0)
    style_sheet(ws_all, all_df)

    # Benchmarks sheet (second tab)
    ws_bm = wb.create_sheet(title="Benchmarks", index=1)
    style_sheet(ws_bm, benchmark_df)

    # Individual category sheets
    for cat, df in category_dfs.items():
        ws = wb.create_sheet(title=cat[:31])
        style_sheet(ws, df)

    wb.save(path)
    print(f"\n✅  Saved: {path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    category_dfs = {}
    all_series   = []

    # ── Fund categories ────────────────────────────────────────────────────────
    for category, fund_list in FUNDS.items():
        print(f"\n📂  {category}")
        df = build_df_from_list(fund_list)
        category_dfs[category] = df
        for col in df.columns:
            all_series.append(df[col])

    # ── Benchmark indices ──────────────────────────────────────────────────────
    print("\n📊  Benchmarks")
    benchmark_df = build_df_from_list(BENCHMARKS)

    # Also include benchmarks in the "All Funds" combined sheet
    for col in benchmark_df.columns:
        all_series.append(benchmark_df[col])

    # ── Build combined sheet ───────────────────────────────────────────────────
    if all_series:
        all_df = pd.concat(all_series, axis=1)
    else:
        all_df = benchmark_df.copy()

    output_path = "nav_data.xlsx"
    write_excel(category_dfs, benchmark_df, all_df, output_path)


if __name__ == "__main__":
    main()