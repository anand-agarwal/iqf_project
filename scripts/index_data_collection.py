"""
Fetch end-of-month BSE benchmark index data using the BseIndiaApi package.
Source  : bseindia.com (via unofficial `bse` Python wrapper)
Output  : benchmark_bse.xlsx

Install :
    pip install bse openpyxl pandas

Usage   :
    python benchmark_bseindiaapi.py
"""

import time
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bse import BSE
from bse.constants import INDEX

# ── Date range ─────────────────────────────────────────────────────────────────
FROM_DATE  = date(2020, 12, 31)
TO_DATE    = date(2026, 1, 31)

# ── Benchmark indices ──────────────────────────────────────────────────────────
# Keys  : display names used as column headers in Excel
# Values: INDEX constants from bse.constants
#
# Full list of available INDEX constants:
#   https://bennythadikaran.github.io/BseIndiaApi/Constants.html
#
BENCHMARKS = {
    "BSE Teck":            INDEX.TECK,
    "BSE Healthcare":      INDEX.HEALTHCARE,
    "BSE IT":              INDEX.INFORMATION_TECHNOLOGY,
    "BSE FMCG":            INDEX.FAST_MOVING_CONSUMER_GOODS,
    "BSE Bankex":          INDEX.BANKEX,
    "BSE Capital Goods":   INDEX.CAPITAL_GOODS,
    "BSE Sensex":          INDEX.SENSEX,
    "BSE Midcap":          INDEX.MIDCAP,
    "BSE Smallcap":        INDEX.SMALLCAP,
}

DOWNLOAD_DIR = Path("./bse_downloads")   # temp dir for CSVs downloaded by bse

# ── Excel styling ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="EBF3FB")
NORMAL_FONT = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center")
RIGHT       = Alignment(horizontal="right")


# ── Helpers ────────────────────────────────────────────────────────────────────

def fetch_index_eom(bse: BSE, index_name: str, display_name: str) -> pd.Series:
    """
    Download daily index data from BSEIndia, resample to end-of-month,
    and return a named Series trimmed to [FROM_DATE, TO_DATE].
    """
    print(f"  Fetching: {display_name} ({index_name})")
    try:
        csv_path = bse.fetchHistoricalIndexData(
            index     = index_name,
            from_date = FROM_DATE,
            to_date   = TO_DATE,
            period    = "D",          # daily; we'll do EOM resampling ourselves
            folder    = DOWNLOAD_DIR,
        )

        if csv_path is None:
            print(f"  ⚠  Empty response for {display_name}")
            return pd.Series(dtype=float, name=display_name)

        df = pd.read_csv(csv_path)

        # BSEIndia CSV columns: Date, Open, High, Low, Close (names may vary)
        # Normalise column names to lowercase for safety
        df.columns = [c.strip().lower() for c in df.columns]

        # Identify the date column (usually 'date' or 'dttm')
        date_col  = next(c for c in df.columns if "date" in c or "dttm" in c)
        # Identify the close column
        close_col = next(c for c in df.columns if "close" in c)

        df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors="coerce")
        df = df.dropna(subset=[date_col]).set_index(date_col).sort_index()

        close = pd.to_numeric(df[close_col], errors="coerce")

        # Resample to last trading day of each month
        eom = close.resample("ME").last()
        eom.name = display_name
        return eom

    except StopIteration:
        print(f"  ✗  Could not identify date/close column for {display_name}")
        return pd.Series(dtype=float, name=display_name)
    except Exception as exc:
        print(f"  ✗  Error fetching {display_name}: {exc}")
        return pd.Series(dtype=float, name=display_name)


def build_benchmark_df(bse: BSE) -> pd.DataFrame:
    series_list = []
    for display_name, index_const in BENCHMARKS.items():
        s = fetch_index_eom(bse, index_const, display_name)
        series_list.append(s)
        time.sleep(0.5)   # polite rate-limiting

    if not series_list:
        return pd.DataFrame()

    df = pd.concat(series_list, axis=1)
    df.index = pd.to_datetime(df.index, errors="coerce")
    df = df[~df.index.isna()].sort_index()
    df.index = df.index.strftime("%d-%b-%Y")
    df.index.name = "Month End Date"
    return df


# ── Excel writer ───────────────────────────────────────────────────────────────

def style_sheet(ws, df: pd.DataFrame):
    for col_idx, col_name in enumerate(["Month End Date"] + list(df.columns), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    for row_idx, (date_str, row) in enumerate(df.iterrows(), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None

        date_cell = ws.cell(row=row_idx, column=1, value=date_str)
        date_cell.font      = NORMAL_FONT
        date_cell.alignment = CENTER
        if fill:
            date_cell.fill = fill

        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(
                row=row_idx, column=col_idx,
                value=round(float(val), 4) if pd.notna(val) else "N/A",
            )
            cell.font          = NORMAL_FONT
            cell.alignment     = RIGHT
            cell.number_format = "#,##0.0000"
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 16
    for col_idx in range(2, len(df.columns) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28
    ws.freeze_panes = "A2"


def write_excel(df: pd.DataFrame, path: str):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Benchmarks")
    style_sheet(ws, df)
    wb.save(path)
    print(f"\n✅  Saved: {path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    DOWNLOAD_DIR.mkdir(exist_ok=True)

    with BSE(download_folder=DOWNLOAD_DIR) as bse:
        # Optional: print all available index names to help you pick constants
        # print(bse.fetchIndexNames())

        print("📊  Fetching BSE benchmark indices via BseIndiaApi...\n")
        df = build_benchmark_df(bse)

    if df.empty:
        print("❌  No data fetched. Check index names or network connection.")
        return

    print(f"\nFetched {len(df)} monthly rows for {len(df.columns)} indices.")
    print(df.head())

    write_excel(df, "benchmark_bse.xlsx")


if __name__ == "__main__":
    main()